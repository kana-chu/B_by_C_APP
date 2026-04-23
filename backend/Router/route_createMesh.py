"""
@component Mod_Api_Mesh_Create
@folder backend/Router
@category API
@description
    Electron → React → FastAPI フローで送られた
    メッシュ生成パラメータをもとに、
    メッシュ用 Excel ファイルを生成する API
"""

from fastapi import APIRouter, HTTPException, Body, Request
from pydantic import BaseModel
from fastapi import Request
import os
import pandas as pd

from Feature.createMesh.f_cM_createDf import f_cM_createDf
from Feature.createMesh.f_cM_CreateXlsxFile import f_cM_CreateXlsxFile

# --------------------------------------------------
# load 確認用ログ
# --------------------------------------------------
print("🔥🔥🔥 ROUTE_MESHSHEET LOADED 🔥🔥🔥", __file__)

# --------------------------------------------------
# Router 定義
# --------------------------------------------------
router = APIRouter(prefix="/meshsheet", tags=["meshsheet"])


# --------------------------------------------------
# ▼ デバッグ・疎通確認用
# --------------------------------------------------
@router.post("/create-mesh-test")
async def create_mesh_test():
    return {"ok": True}


# ======================================================
# ▼ create-mesh（メイン処理）
# ======================================================
class CreateMeshRequest(BaseModel):
    left_up: str
    right_down: str
    ex_meshSize: str
    save_path: str


@router.post("/create-mesh")
async def create_mesh(
    req: CreateMeshRequest = Body(...),
    request: Request = None,
):
    logs: list[str] = []

    def log(msg: str):
        print(msg)
        logs.append(msg)

    """
    Mesh シート生成 API（JSON版・完成形）
    Python 側で DataFrame 生成〜Excel 書き込みまでを完結させる
    """

    # --------------------------------------------------
    # ▼ 受信デバッグ（開発用）
    # --------------------------------------------------
    if request:
        try:
            log(f"🔥 HEADERS: {dict(request.headers)}")
            log(f"🔥 RAW JSON: {await request.json()}")
        except Exception as e:
            log(f"🔥 RAW JSON READ ERROR: {e}")

    log(f"🔥 PARSED MODEL: {req}")

    # --------------------------------------------------
    # 0. 入力値チェック
    # --------------------------------------------------
    if len(req.left_up) != 8 or len(req.right_down) != 8:
        raise HTTPException(
            status_code=400,
            detail="left_up / right_down は8桁メッシュコードである必要があります",
        )


    save_path = req.save_path

    # 親ディレクトリを取得
    save_dir = os.path.dirname(save_path)

    # 親ディレクトリの存在チェック
    if not os.path.isdir(save_dir):
        raise HTTPException(
            status_code=400,
            detail=f"保存先ディレクトリが存在しません: {save_dir}",
        )


    # --------------------------------------------------
    # 1. Excel ファイル新規作成
    # --------------------------------------------------
    try:
        save_filepath = f_cM_CreateXlsxFile(req.save_path)
        log(f"✅ Excel created: {save_path}")
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Excel ファイル作成に失敗しました: {e}",
        )

    # --------------------------------------------------
    # 2. A1～D4 に left_up の上4桁を書き込み
    # --------------------------------------------------
    from openpyxl import load_workbook

    try:
        mcode4 = str(req.left_up)[:4]

        wb = load_workbook(save_path)
        ws = wb.active

        for row in range(1, 5):        # A1～D4
            for col in range(1, 5):
                ws.cell(row=row, column=col).value = mcode4

        wb.save(save_path)
        log(f"✅ A1:D4 に 1次地域メッシュコード設定: {mcode4}")

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"A1～D4 メッシュコード書き込みに失敗しました: {e}",
        )


    # --------------------------------------------------
    # 3. lon / lat DataFrame 生成
    # --------------------------------------------------
    try:
        df_lon = f_cM_createDf(
            req.left_up,
            req.right_down,
            req.ex_meshSize,
            lonlat="lon",
        )

        df_lat = f_cM_createDf(
            req.left_up,
            req.right_down,
            req.ex_meshSize,
            lonlat="lat",
        )

        log("✅ DataFrame generated (lon / lat)")
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"メッシュ DataFrame 生成に失敗しました: {e}",
        )

    # --------------------------------------------------
    # 4. Excel へ書き込み
    # --------------------------------------------------
    try:
        with pd.ExcelWriter(
            save_filepath,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="overlay",
        ) as writer:

            sheet_name = writer.book.sheetnames[0]

            # lon（縦）
            df_lon.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=False,
                startrow=4,
            )

            # lat（横）
            df_lat.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=False,
                startrow=0,
                startcol=4,
            )

        log("✅ メッシュシートファイル作成完了")
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Excel 書き込みに失敗しました: {e}",
        )

    # --------------------------------------------------
    # 5. レスポンス
    # --------------------------------------------------
    return {
        "message": "OK",
        "params": {
            "left_up": req.left_up,
            "right_down": req.right_down,
            "ex_meshSize": req.ex_meshSize,
        },
        "save_path": save_filepath,
        "logs": logs,
    }
