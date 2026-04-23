"""
@component f_cC_readItemList
@folder backend/Feature/calcCensus
@category DataProcessing
@description
    絶対パスで指定された Excel ファイルから、
    指定された「項目シート」を読み込み、
    A/B/C 列（項目名・列名・セル名）を
    行単位のリストとして返す。
    ※ 1行目はヘッダーとして除外する
"""

import os
from openpyxl import load_workbook


def f_cC_readItemList(input_file: str, sheet_name: str):
    print("f_cC_readItemList 開始")

    # --------------------------------------------------
    # 入力チェック
    # --------------------------------------------------
    if not os.path.isfile(input_file):
        raise FileNotFoundError(
            f"Excel ファイルが存在しません: {input_file}"
        )

    if not input_file.lower().endswith(".xlsx"):
        raise ValueError(
            "対応しているのは .xlsx ファイルのみです"
        )

    # --------------------------------------------------
    # Workbook 読み込み（read_only）
    # --------------------------------------------------
    wb = load_workbook(
        filename=input_file,
        read_only=True,
        data_only=True,
    )

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"指定されたシートが存在しません: {sheet_name}"
        )

    ws = wb[sheet_name]

    item_list = []

    # --------------------------------------------------
    # A/B/C 列を走査（★1行目はヘッダーなので除外）
    # --------------------------------------------------
    for row in ws.iter_rows(min_row=2, values_only=True):
        sheet_val = row[0] if len(row) > 0 else None
        col_val   = row[1] if len(row) > 1 else None
        head_val  = row[2] if len(row) > 2 else None

        # 完全空行はスキップ
        if not sheet_val and not col_val and not head_val:
            continue

        item_list.append({
            "sheetName": str(sheet_val) if sheet_val is not None else "",
            "columnIndex": str(col_val) if col_val is not None else "",
            "headerName": str(head_val) if head_val is not None else "",
        })

    wb.close()

    print("f_cC_readItemList 終了")

    return {
        "itemList": item_list
    }
