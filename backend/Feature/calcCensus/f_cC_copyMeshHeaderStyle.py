from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from copy import copy


def add_border(cell, *, left=None, right=None, top=None, bottom=None):
    b = cell.border
    cell.border = Border(
        left=left if left else b.left,
        right=right if right else b.right,
        top=top if top else b.top,
        bottom=bottom if bottom else b.bottom,
    )


def f_cC_copyMeshHeaderStyle(
    file_path: str,
    item_sheetNames: list[str],
    exMeshSize: int,
    source_sheetName: str = "メッシュ",
):
    ROW_BLOCK = {5: 200, 25: 40, 50: 20}
    COL_BLOCK = {5: 200, 25: 40, 50: 20}

    row_block = ROW_BLOCK[exMeshSize]
    col_block = COL_BLOCK[exMeshSize]

    wb = load_workbook(file_path)
    ws_src = wb[source_sheetName]

    thin = Side(style="thin")
    center = Alignment(horizontal="center", vertical="center")

    for sheet_name in item_sheetNames:
        ws = wb[sheet_name]

        # ==================================================
        # 0. 既存結合解除
        # ==================================================
        for merged in list(ws.merged_cells.ranges):
            ws.unmerge_cells(merged.coord)

        max_row = ws.max_row
        max_col = ws.max_column

        # ==================================================
        # 1. A1:D4 固定結合
        # ==================================================
        ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=4)
        src = ws_src["A1"]
        dst = ws["A1"]
        dst.value = src.value
        dst.alignment = center

        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)

        # ==================================================
        # 2. A,B列 縦結合 + 境界罫線
        # ==================================================
        for col in (1, 2):
            for start in range(5, max_row + 1, row_block):
                end = min(start + row_block - 1, max_row)
                ws.merge_cells(start_row=start, start_column=col,
                               end_row=end, end_column=col)
                ws.cell(row=start, column=col).alignment = center

                for c in range(1, max_col + 1):
                    add_border(ws.cell(row=end, column=c), bottom=thin)

        # ==================================================
        # 3. 1,2行 横結合 + 境界罫線
        # ==================================================
        for row in (1, 2):
            for start in range(5, max_col + 1, col_block):
                end = min(start + col_block - 1, max_col)
                ws.merge_cells(start_row=row, start_column=start,
                               end_row=row, end_column=end)
                ws.cell(row=row, column=start).alignment = center

                for r in range(1, max_row + 1):
                    add_border(ws.cell(row=r, column=end), right=thin)

        # ==================================================
        # 4. 4-5行目境界線
        # ==================================================
        for c in range(1, max_col + 1):
            add_border(ws.cell(row=4, column=c), bottom=thin)

        # ==================================================
        # 5. A,B,C列 右罫線（A1:D4除外）
        # ==================================================
        for col in (1, 2, 3):
            for r in range(5, max_row + 1):
                add_border(ws.cell(row=r, column=col), right=thin)

        # ==================================================
        # ✅ 6. D列右罫線を必ず付与（今回の追加）
        # ==================================================
        for r in range(1, max_row + 1):
            add_border(ws.cell(row=r, column=4), right=thin)

        # ==================================================
        # 7. 1,2,3行 下罫線
        # ==================================================
        for row in (1, 2, 3):
            for c in range(1, max_col + 1):
                add_border(ws.cell(row=row, column=c), bottom=thin)

        # ==================================================
        # ✅ 8. 列幅を 3 に設定（今回の追加）
        # ==================================================
        for c in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = 4

    wb.save(file_path)
    wb.close()