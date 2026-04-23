from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


def f_cC_fillZeroCellsGray(
    file_path: str,
    sheet_names: list[str],
):
    """
    データ領域（5行目以降 × E列以降）に対して：
    ・値が 0 → 背景を薄グレー
    ・空白セル → 「0」で埋め、文字色薄グレー＋背景薄グレー
    """

    wb = load_workbook(file_path)

    # 薄グレー（背景）
    gray_fill = PatternFill(
        start_color="DDDDDD",
        end_color="DDDDDD",
        fill_type="solid",
    )

    # 薄グレー（文字）
    gray_font = Font(color="888888")

    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        for row in range(5, ws.max_row + 1):          # 5行目以降
            for col in range(5, ws.max_column + 1):   # E列以降
                cell = ws.cell(row=row, column=col)

                # 結合セルの子セルはスキップ
                if cell.coordinate in ws.merged_cells:
                    continue

                val = cell.value

                # --- 空白セル ---
                if val is None or val == "":
                    cell.value = 0
                    cell.fill = gray_fill
                    cell.font = gray_font

                # --- 数値の 0 ---
                elif isinstance(val, (int, float)) and val == 0:
                    cell.fill = gray_fill
                    # フォントは既存のまま

    wb.save(file_path)
    wb.close()