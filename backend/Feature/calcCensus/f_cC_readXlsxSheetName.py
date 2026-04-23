"""
@component f_cC_readXlsxSheetName
@folder backend/Feature
@category DataProcessing
@description
    絶対パスで指定された Excel ファイルを読み込み、
    含まれる全シート名を取得して返す。

@params
    input_file : str
        読み込む Excel ファイルの絶対パス（.xlsx）

@return
    dict {
        "sheet_names": List[str]
    }

@usage
    result = f_cC_readXlsxSheetName("C:/path/to/file.xlsx")
    names = result["sheet_names"]

@remarks
    - openpyxl を使用
    - 読み取り専用（Excel の内容は変更しない）
"""

import os
from openpyxl import load_workbook


def f_cC_readXlsxSheetName(input_file: str):
    print("f_cC_readXlsxSheetName 開始")

    # ▼ パスチェック
    if not os.path.isfile(input_file):
        raise FileNotFoundError(
            f"Excel ファイルが存在しません: {input_file}"
        )

    # ▼ 拡張子チェック
    if not input_file.lower().endswith(".xlsx"):
        raise ValueError(
            "対応しているのは .xlsx ファイルのみです"
        )

    # ▼ Workbook 読み込み（read_only）
    wb = load_workbook(
        filename=input_file,
        read_only=True,
        data_only=True,
    )

    sheet_names = wb.sheetnames

    wb.close()

    print("f_cC_readXlsxSheetName 終了")

    return {
        "sheet_names": sheet_names
    }