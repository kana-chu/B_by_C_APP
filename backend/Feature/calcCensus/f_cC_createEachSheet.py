"""
@component f_cC_createEachSheet
@folder backend/Feature/calcCensus
@category DataProcessing
@description
    国勢調査・経済センサス用 Excel ファイルに対して、
    「メッシュ」シートを雛形として、
    itemList に含まれる各 sheetName ごとに
    新しいシートを作成する。

@params
    file_path : str
        対象 Excel ファイルの絶対パス
    item_list : list[dict]
        f_cC_readItemList で取得した項目定義リスト

@return
    dict {
        "createdSheets": list[str]
    }

@usage
    result = f_cC_createEachSheet(file_path, item_list)
    created = result["createdSheets"]

@remarks
    - openpyxl を使用
    - 雛形シート名は「メッシュ」固定
    - 既存シートがある場合はスキップ
"""

from openpyxl import load_workbook
import os


def f_cC_createEachSheet(file_path: str, item_list: list):
    print("f_cC_createEachSheet 開始")

    # --------------------------------------------------
    # 入力チェック
    # --------------------------------------------------
    if not os.path.isfile(file_path):
        raise FileNotFoundError(
            f"Excel ファイルが存在しません: {file_path}"
        )

    if not isinstance(item_list, list):
        raise ValueError("item_list は list である必要があります")

    # --------------------------------------------------
    # Workbook 読み込み
    # --------------------------------------------------
    wb = load_workbook(file_path)

    source_sheet_name = "メッシュ"

    if source_sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"'{source_sheet_name}' シートが見つかりません"
        )

    source_sheet = wb[source_sheet_name]

    created_sheets = []
    seen_names = set()

    # --------------------------------------------------
    # item_list から sheetName を抽出してシート作成
    # --------------------------------------------------
    for item in item_list:
        sheet_name = item.get("sheetName")

        # 無効データはスキップ
        if not sheet_name:
            continue

        # 同一 sheetName の重複作成防止
        if sheet_name in seen_names:
            continue
        seen_names.add(sheet_name)

        if sheet_name in wb.sheetnames:
            print(f"ℹ️ シート '{sheet_name}' は既に存在 → スキップ")
            continue

        new_sheet = wb.copy_worksheet(source_sheet)
        new_sheet.title = sheet_name
        created_sheets.append(sheet_name)

    # --------------------------------------------------
    # 保存
    # --------------------------------------------------
    wb.save(file_path)
    wb.close()

    print(f"✅ {len(created_sheets)} シート作成完了")

    return {
        "createdSheets": created_sheets
    }